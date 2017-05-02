import unirest, json

"""
# These code snippets use an open-source library. http://unirest.io/python

print(response.raw_body)

print(response.body['pronunciation']['all'])


result=json.loads({"word":"resume","pronunciation":{"all":"r'zum"}})
print(result['pronunciation']['all'])
"""

from openpyxl import Workbook
from openpyxl import load_workbook

import warnings

warnings.filterwarnings("ignore")


def get_pronunciation(word):
    url = "https://wordsapiv1.p.mashape.com/words/" + word + "/pronunciation"
    print(url)
    try:
        response = unirest.get(url,
                               headers={
                                   "X-Mashape-Key": "Rxx96iAmPtmsh7QzQ18RyvTyIKLop1tN4Gzjsn99lrgaYA1wq2",
                                   "Accept": "application/json"
                               }
                               )
    except:
        print ("check word: " + word)
        response = unirest.get("https://jchencha-autosuggest.p.mashape.com/suggest/?word=" + word,
                               headers={
                                   "X-Mashape-Key": "Rxx96iAmPtmsh7QzQ18RyvTyIKLop1tN4Gzjsn99lrgaYA1wq2",
                                   "Accept": "application/json"
                               }
                               )
        return get_pronunciation(response.body['mispelt']['suggestions'][0])
    else:
        print(response.raw_body)
        if 'pronunciation' in response.body:
            result = response.body['pronunciation']
            if 'all' in result:
                return result['all']
            if 'verb' in result:
                return result['verb']
            return result
        else:
            return "**"


wb = load_workbook(filename='norang.xlsx')
ws = wb.active  # ws = wb['Sheet1']

count = 0

try:
    for i in range(2, ws.get_highest_row() + 1):
        if ws['D' + str(i)].value is None:
            ws['D' + str(i)].value = get_pronunciation(ws['B' + str(i)].value)
            print(ws['B' + str(i)].value)
            count += 1
            if count > 10:
                wb.save('norang.xlsx')
                print('**saved')
                count = 0

                # ws['B'+str(1190)].value= get_pronunciation(ws['B'+str(1190)].value)
finally:
    wb.save('norang.xlsx')
