import scrapy



class BlogSpider(scrapy.Spider):
    name = 'blogspider'

    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb = load_workbook(filename='norang.xlsx')
    ws = wb.active  # ws = wb['Sheet1']

    count = 0
    start_urls=[]
    try:
        for i in range(2, ws.get_highest_row() + 1):
            if ws['D' + str(i)].value is None:
                ws['D' + str(i)].value = "checked"
                print(ws['B' + str(i)].value)
                start_urls.append('http://aha-dic.com/View.asp?word='+ws['B' + str(i)].value)
                count += 1
                if count > 100:
                    wb.save('norang.xlsx')
                    print('**saved')
                    break #count = 0
    finally:
        wb.save('norang.xlsx')




    #start_urls = ['http://aha-dic.com/View.asp?word=distributuion'#,
    #              'http://aha-dic.com/View.asp?word=question'
    #]

    def parse(self, response):
        print (response.css('.word').extract())
        if response.css('.word').extract() == []:
            yield scrapy.Request("http://aha-dic.com/"+response.css('#container_result a::attr(href)').extract()[0], self.parse)
        else:
            yield {'star':response.css('.star').extract(), 'word':response.css('.word').extract(), 'phoneticKor':response.css('.phoneticKor').extract(), 'phonetic':response.css('.phonetic').extract()}

        #for url in response.css('ul li a::attr("href")').re('.*/category/.*'):
            
            #yield scrapy.Request(response.urljoin(url), self.parse_titles)

    def parse_titles(self, response):
        for post_title in response.css('div.entries > ul > li a::text').extract():
            yield {'title': post_title}
