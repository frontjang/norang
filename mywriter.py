# -*- coding: utf-8 -*-

import scrapy
from scrapy import signals
from scrapy.xlib.pydispatch import dispatcher
import re

import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

currentRow=1



class BlogSpider(scrapy.Spider):
    name = 'blogspider'
    start_urls=[]
    custom_settings = {
        'ITEM_PIPELINES': {
            'mywriter.BlogSpider': 400
        }
    }
    
    url=None
    
    
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb = load_workbook(filename='all.xlsx')
    ws = wb.active  # ws = wb['Sheet1']

    import xlsxwriter
    workbook = xlsxwriter.Workbook('output.xlsx')
    worksheet = workbook.add_worksheet()
    highlight = workbook.add_format({'underline': True, 'bold': True})
    
    def __init__(self):
        #dispatcher.connect(self.spider_closed, signals.spider_closed)
        dispatcher.connect(self.engine_stopped, signals.engine_stopped)
        print("    def __init__(self):")
        
        self.url=self.next_url()
    
    def spider_closed(self, spider, reason):
        print("spider_closed")
        
    def engine_stopped(self):
        self.wb.save('all.xlsx')
        self.workbook.close()
        print("engine_closed")
        

    def next_url(self):
        global currentRow
        #print("next_url:"+self.ws['E' + str(currentRow)].value)
        #    print("************currentRow1:"+str(self))
        #    print("************currentRow1:"+str(currentRow))
        while True:
            while(self.ws['E' + str(currentRow)].value == "failed" or self.ws['E' + str(currentRow)].value == "checked"):
                currentRow=currentRow+1

            yield 'http://aha-dic.com/View.asp?word='+self.ws['C' + str(currentRow)].value

    
    
    def start_requests(self):
        while True:
            yield scrapy.Request(self.url.next())

    def parse(self, response):
        print (response.css('.word').extract())
        global currentRow
        if response.css('.word').extract() == []:
            try:
                correct = response.css('#container_result > div:nth-child(2) a::attr(href)').extract()[0]
                yield scrapy.Request("http://aha-dic.com/"+correct, self.parse)
            except:
                self.ws['E' + str(currentRow)].value = "failed"
                currentRow=currentRow+1
                self.start_requests()
        else:
            yield {'star':response.css('.star').extract(), 'word':response.css('.word').extract(), 'phoneticKor':response.css('.phoneticKor').extract(), 'phonetic':response.css('.phonetic').extract()}


    def process_item(self, item, spider):
        global currentRow
        currentRow=currentRow+1
        realCurrentRow=currentRow-1;
        adjustedCurrentRow=realCurrentRow-1
        print("************currentRow2:"+str(self))
        print("************currentRow2:"+str(currentRow))
        self.ws['E' + str(realCurrentRow)].value = "checked"

        if item['word'] != []:
            self.worksheet.write(adjustedCurrentRow, 0, self.ws['A'+str(realCurrentRow)].value) #day
            self.worksheet.write(adjustedCurrentRow, 1, self.ws['B'+str(realCurrentRow)].value) #num
            self.worksheet.write(adjustedCurrentRow, 2, self.ws['C'+str(realCurrentRow)].value) #word
            self.worksheet.write(adjustedCurrentRow, 3, self.ws['D'+str(realCurrentRow)].value) #kor


            p = re.compile("<\/span>(.+)<span", re.MULTILINE|re.DOTALL)
            self.worksheet.write(adjustedCurrentRow, 4, re.search(p, item['phonetic'][0]).group(1).strip()) #pho

            p = re.compile("\[(.+)\]", re.MULTILINE|re.DOTALL)
            phoKor=re.search(p, item['phoneticKor'][0]).group(1);

            phoKor=phoKor.replace('<span class="accent">', "', self.highlight, '").replace('</span>', "', '")
            phoKorStr="self.worksheet.write_rich_string('F"+str(realCurrentRow)+"', '"+phoKor+"')"
            print (phoKorStr)
            exec(phoKorStr)            

            """self.worksheet.write(row, 'B'+row, self.ws['B'+row].value)
            self.worksheet.write('C'+row, self.ws['C'+row].value)
            self.worksheet.write('D'+row, self.ws['D'+row].value)
            phonetic=item['phonetic']
            p = re.compile(ur'\/span>.+\\t(\S+?)\\r.+<span')
            print(re.match(p, item['phonetic']))

  
            self.worksheet.write_rich_string('E'+currentRow, item['phonetic'])
            self.worksheet.write_rich_string('F'+currentRow, item['phoneticKor'])
            self.ws['D' + row].value = "checked"
"""
            
            return item
            
        else:
            raise DropItem("Missing price in %s" % item)
